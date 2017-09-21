#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
Looks subfolders for .xlsx files.
Using openpyxl library removes rows with technical data
from Excel worksheets and saves to new file.
"""

__author__ = "Roman Ivanov"
__copyright__ = "Copyright (C) 2017 Roman Ivanov"
__license__ = "Public Domain"
__version__ = "1.0"

import sys
from os import path, listdir, sep
from openpyxl import load_workbook, Workbook 

d = "."
ext = ".xlsx"

subdirs = [path.join(d, o) for o in listdir(d) 
                    if path.isdir(path.join(d,o))]

def main():
  for sd in subdirs:
    if path.split(sd)[1][0] == d:
      continue
    printCP(sd)
    processDirectory(sd)
  raw_input("Press Enter to continue...")

def printCP(s):
  if sys.stdout.encoding != "UTF-8":
    print s.decode('cp1251')
  else:
    print s

def processDirectory(d):
  xslxFiles = [path.join(d, o) for o in listdir(d) 
                    if path.splitext(path.join(d,o))[1].lower() == ext]
  wb = Workbook(write_only=True)
  isEmpty = True
  for xf in xslxFiles:
    sls = processSrtFile(xf)
    for sl in sls:
      isEmpty = False
      ws = wb.create_sheet(title=path.splitext(path.split(xf)[1])[0].decode("UTF-8"))
      for row in sl:
        ws.append(row)
  filename = d + sep + "!Dialogs" + ext
  if not isEmpty:
    wb.save(filename)

def processSrtFile(filename):
  printCP("\t" + path.split(filename)[1])
  wb = load_workbook(filename, data_only=True)
  names = wb.sheetnames
  sls = []
  for name in names:
    ws = wb.get_sheet_by_name(name)
    sl = processSheet(ws)
    if sl != []:
      sls.append(sl)
  return sls

def processSheet(ws):
    subtitle_id = 1
    text_row_in = 0
    subtitle_list = []
    for row in ws.iter_rows(max_col=2):
      subtitle_row = []
      
      if text_row_in != 0:
        text_row_in -= 1
      cell_coord = 1
      
      for cell in row:
        if text_row_in == 1:
          subtitle_row.append(cell.value)
        else:
          if cell.value == subtitle_id:
            subtitle_id += 1
            text_row_in = 3

      if subtitle_row != []:
        subtitle_list.append(subtitle_row)
    return subtitle_list

def writeResultFile(filename, speechlines):
  wb = Workbook(write_only=True)
  ws = wb.create_sheet()
  for row in speechlines:
    ws.append(row)
  wb.save(filename)

if __name__ == "__main__":
    main()
